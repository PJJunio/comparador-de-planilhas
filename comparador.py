import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from openpyxl import *
from openpyxl.styles import *;
from openpyxl.utils import get_column_letter

def comparar():

    root = tk.Tk()
    root.withdraw()

    file_types = ([("Arquivos Excel (.xlsx)", "*.xlsx"), ("Todos os arquivos", "*.*")])

    messagebox.showinfo("Seleção", "Selecione a PRIMEIRA planilha")
    plan1 = filedialog.askopenfilename(
        title="Seleciona uma planilha",
        filetypes=file_types
    )
    if not plan1:
        return
    
    messagebox.showinfo("Seleção", "Selecione a SEGUNDA planilha")
    plan2 = filedialog.askopenfilename(
        title="Seleciona uma planilha",
        filetypes=file_types
    )
    if not plan2:
        return
    
    messagebox.showinfo("Seleção", "Selecione onde salvar o resultado")
    plansaida = filedialog.asksaveasfilename(
        title="Salvar como",
        defaultextension=".xlsx",
        filetypes=file_types,
        initialfile="resultado.xlsx"
    )
    if not plansaida:
        return
    
    try:
        wb1 = load_workbook(plan1)
        wb2 = load_workbook(plan2)

        wb_saida = Workbook()
        ws_saida = wb_saida.active
        ws_saida.title = "comparacao"

        fill_amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        borda = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))

        ws1 = wb1.active
        ws2 = wb2.active

        max_row = max(ws1.max_row, ws2.max_row)

        for row in range (1, max_row + 1):

            valor1_1 = ws1.cell(row=row, column=1).value
            valor1_2 = ws1.cell(row=row, column=2).value
            valor1_3 = ws1.cell(row=row, column=3).value
            valor1_4 = ws1.cell(row=row, column=4).value
            valor1_5 = ws1.cell(row=row, column=5).value

            valor2_1 = ws2.cell(row=row, column=1).value
            valor2_2 = ws2.cell(row=row, column=2).value
            valor2_3 = ws2.cell(row=row, column=3).value
            valor2_4 = ws2.cell(row=row, column=4).value
            valor2_5 = ws2.cell(row=row, column=5).value

            ws_saida.cell(row=row, column=1, value=valor1_1)
            ws_saida.cell(row=row, column=2, value=valor1_2)
            ws_saida.cell(row=row, column=3, value=valor1_3)
            ws_saida.cell(row=row, column=4, value=valor1_4)
            ws_saida.cell(row=row, column=5, value=valor1_5)

            ws_saida.cell(row=row, column=7, value=valor2_1)
            ws_saida.cell(row=row, column=8, value=valor2_2)
            ws_saida.cell(row=row, column=9, value=valor2_3)
            ws_saida.cell(row=row, column=10, value=valor2_4)
            ws_saida.cell(row=row, column=11, value=valor2_5)

            for col in range(1, 12):
                celula = ws_saida.cell(row=row, column=col)
                if celula.value is not None:
                    celula.border = borda
            
            if valor1_1 == valor2_1 and valor1_5 == valor2_5:
                for col in [1, 2, 3, 4, 5, 7, 8, 9, 10, 11]:
                    celula = ws_saida.cell(row=row, column=col)
                    celula.fill = fill_amarelo

        for col in ws_saida.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = (max_length + 2) * 1.2
            ws_saida.column_dimensions[column].width = adjusted_width

        wb_saida.save(plansaida)
        messagebox.showinfo("Sucesso! resultado salvo em ", plansaida)

    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao processar:\n{str(e)}")
    

if __name__ == "__main__":
    comparar()